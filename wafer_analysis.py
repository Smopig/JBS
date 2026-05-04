#最終版
#1.合併並格式化所有 Excel 原始資料。
#2.為整片 wafer 繪製熱區圖並標記統計資訊。
#3.對整片 wafer 資料按 Lot 分組並進行數值統計輸出。

import os
import time
import pandas as pd
import numpy as np
from scipy import stats
from tkinter import Tk, simpledialog, filedialog
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, LineChart, Reference

# 設定樣式
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

# ==========================================
# Part 1: 合併與格式化 Excel
# ==========================================

def get_lot_name(filename):
    # AAA002Y-N64S03A_01-data.xls → N64S03A_01
    base_name = os.path.splitext(filename)[0]
    parts = base_name.split('-')
    if len(parts) >= 2:
        return parts[1]
    return base_name

def format_excel_file(filepath):
    wb = load_workbook(filepath)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            cell.font = Font(name='Calibri')
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if isinstance(cell.value, (int, float)) and cell.value is not None:
                val = cell.value
                if isinstance(val, float):
                    decimal_str = str(val).split('.')[-1] if '.' in str(val) else ''
                    if len(decimal_str) > 4:
                        cell.number_format = '0.00E+00'
                    elif val == int(val):
                        cell.number_format = '0'
                    else:
                        cell.number_format = 'General'
                elif isinstance(val, int):
                    cell.number_format = '0'

            if cell.value is not None:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filepath)
    print(f"🧾 Excel 格式設定完成：{filepath}")

def merge_excel_with_custom_header(folder_path, output_folder, header_row_index=12):
    all_data = []
    error_log = []

    for filename in os.listdir(folder_path):
        if filename.endswith(('.xls', '.xlsx')):
            file_path = os.path.join(folder_path, filename)
            try:
                df = pd.read_excel(file_path, header=None)

                if df.shape[0] <= header_row_index:
                    error_log.append([filename, f"少於 {header_row_index+1} 列，無法取得欄位"])
                    continue

                header_row = df.iloc[header_row_index].tolist()
                if all(pd.isna(h) for h in header_row):
                    error_log.append([filename, "欄位名稱為空"])
                    continue

                if df.shape[0] <= 24:
                    error_log.append([filename, "資料不足25列"])
                    continue

                extracted_rows = df.iloc[29:].copy().dropna(how='all')
                if extracted_rows.empty:
                    error_log.append([filename, "無有效資料"])
                    continue

                actual_cols = extracted_rows.shape[1]
                expected_cols = len(header_row)
                if actual_cols < expected_cols:
                    header_row = header_row[:actual_cols]
                elif actual_cols > expected_cols:
                    header_row += [f"未知欄位_{i}" for i in range(expected_cols, actual_cols)]

                extracted_rows.columns = header_row
                extracted_rows = extracted_rows.loc[:, ~extracted_rows.columns.astype(str).str.contains('SOT_SOT', na=False)]
                extracted_rows = extracted_rows.dropna(axis=1, how='all')
                lot_value = get_lot_name(filename)
                extracted_rows.insert(0, "Lot", lot_value)
                all_data.append(extracted_rows)

            except Exception as e:
                error_log.append([filename, f"讀取失敗: {e}"])

    if not all_data:
        print("🚫 無任何有效資料，結束")
        return None

    merged_df = pd.concat(all_data, ignore_index=True)

    col_rename_dict = {
        '2 VFBC': 'VF_5mA(V)',
        '3 ICBO': 'IR_1V(uA)',
        '4 ICBO2': 'IR_10V(uA)',
        '5 ICBO3': 'IR_30V(uA)',
        '6 ICBO4': 'IR_50V(uA)',
        '7 ICBO5': 'IR_60V(uA)',
        '8 ICBO6': 'IR_68V(uA)',
        '9 ICBO7': 'IR_71V(uA)',
        '10 ICBO8': 'IR_74V(uA)',
        '11 BVCBO': 'VR_10uA(V)',
        '12 BVCBO2': 'VR_100uA(V)',
        '13 BVCBO3': 'VR_1mA(V)',
        '14 BVCBO4': 'VR_10mA(V)',
        '16 VFBC2': 'VF_10mA(V)',
        '17 VFBC3': 'VF_260mA(V)',
        '18 VFBC4': 'VF_1A(V)',
        '19 VFBCH': 'VF_10A(V)'
    }
    merged_df.rename(columns=col_rename_dict, inplace=True)

    desired_order = ["Lot", "Serial", "Bin", "Xpos", "Ypos", "YY,XX",
                     'IR_1V(uA)',
                     'IR_10V(uA)',
                     'IR_30V(uA)',
                     'IR_50V(uA)',
                     'IR_60V(uA)',
                     'IR_68V(uA)',
                     'IR_71V(uA)',
                     'IR_74V(uA)',
                     'VR_10uA(V)',
                     'VR_100uA(V)',
                     'VR_1mA(V)',
                     'VR_10mA(V)',
                     'VF_10mA(V)',
                     'VF_260mA(V)',
                     'VF_1A(V)',
                     'VF_10A(V)']

    existing_cols = [col for col in desired_order if col in merged_df.columns]
    merged_df = merged_df[existing_cols]

    # IR: A → uA (×1e6)，取小數點第2位
    for col in [c for c in merged_df.columns if c.startswith('IR_')]:
        merged_df[col] = pd.to_numeric(merged_df[col], errors='coerce').mul(1e6).round(2)

    # VR: 取小數點第1位
    for col in [c for c in merged_df.columns if c.startswith('VR_')]:
        merged_df[col] = pd.to_numeric(merged_df[col], errors='coerce').round(1)

    # VF: 取小數點第3位
    for col in [c for c in merged_df.columns if c.startswith('VF_')]:
        merged_df[col] = pd.to_numeric(merged_df[col], errors='coerce').round(3)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    merged_output = os.path.join(output_folder, f"合併資料_{timestamp}.xlsx")
    merged_df.to_excel(merged_output, index=False)
    format_excel_file(merged_output)

    return merged_output

# ==========================================
# Part 2: MAP 熱區圖 + 統計（整片 wafer）
# ==========================================

def build_excel_dist_charts(wb, lot_groups, value_columns):
    """
    用 Excel 原生圖表建立常態分布圖。
    排版：同一片 wafer 的所有量測欄位圖表放在同一列。
    資料存於隱藏的「分布圖_data」sheet，圖表顯示於「分布圖」sheet。
    """
    N_BINS = 25
    ROWS_PER_DATASET = N_BINS + 2   # header + N_BINS 資料列 + 1 空白列
    CHART_W = 11    # 圖表寬度（cm）
    CHART_H = 7     # 圖表高度（cm）
    COLS_PER_CHART = 9   # 每張圖水平佔用的欄數（間距控制）
    ROWS_PER_CHART = 15  # 每張圖垂直佔用的列數（間距控制）

    ws_data = wb.create_sheet(title="分布圖_data")
    ws_chart = wb.create_sheet(title="分布圖")

    # 將 lot_groups 展平成有序清單，每個元素 = 一片 wafer
    all_lots = []
    for main_lot, sublots in lot_groups.items():
        for sub_lot, rows in sublots.items():
            all_lots.append((main_lot, sub_lot, rows))

    data_row = 1  # ws_data 目前寫入位置

    for lot_idx, (main_lot, sub_lot, rows) in enumerate(all_lots):
        lot_df = pd.DataFrame(rows)
        lot_label = f"{main_lot}_{sub_lot}" if sub_lot else main_lot
        chart_anchor_row = lot_idx * ROWS_PER_CHART + 1  # 同一 wafer 同一列

        val_cols_present = [c for c in value_columns if c in lot_df.columns]
        for val_idx, val_col in enumerate(val_cols_present):
            series_data = pd.to_numeric(lot_df[val_col], errors='coerce').dropna()
            if len(series_data) < 3:
                data_row += ROWS_PER_DATASET
                continue

            mu, sigma = stats.norm.fit(series_data)
            counts, bin_edges = np.histogram(series_data, bins=N_BINS, density=True)
            bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
            pdf_vals = stats.norm.pdf(bin_centers, mu, sigma)

            # 寫入 header
            ws_data.cell(row=data_row, column=1, value=f"{lot_label}_{val_col}")
            ws_data.cell(row=data_row, column=2, value="Density")
            ws_data.cell(row=data_row, column=3, value="Normal PDF")

            # 寫入資料列
            for i in range(N_BINS):
                r = data_row + 1 + i
                ws_data.cell(row=r, column=1, value=round(float(bin_centers[i]), 4))
                ws_data.cell(row=r, column=2, value=round(float(counts[i]), 6))
                ws_data.cell(row=r, column=3, value=round(float(pdf_vals[i]), 6))

            # 直方圖（BarChart）
            bar = BarChart()
            bar.type = "col"
            bar.grouping = "clustered"
            bar.overlap = 100
            bar.gapWidth = 0
            bar.title = f"{lot_label} | {val_col}  (n={len(series_data)}, mean={mu:.4g}, std={sigma:.4g})"
            bar.y_axis.title = "Density"
            bar.width = CHART_W
            bar.height = CHART_H

            hist_ref = Reference(ws_data, min_col=2, min_row=data_row, max_row=data_row + N_BINS)
            cats = Reference(ws_data, min_col=1, min_row=data_row + 1, max_row=data_row + N_BINS)
            bar.add_data(hist_ref, titles_from_data=True)
            bar.set_categories(cats)

            # 常態分布曲線（LineChart）
            line = LineChart()
            pdf_ref = Reference(ws_data, min_col=3, min_row=data_row, max_row=data_row + N_BINS)
            line.add_data(pdf_ref, titles_from_data=True)
            for s in line.series:
                s.smooth = True

            bar += line  # 合併成組合圖

            chart_anchor_col = val_idx * COLS_PER_CHART + 1
            anchor = ws_chart.cell(row=chart_anchor_row, column=chart_anchor_col).coordinate
            ws_chart.add_chart(bar, anchor)

            data_row += ROWS_PER_DATASET

    ws_data.sheet_state = 'hidden'


def build_pivot(df, val_col):
    df[val_col] = pd.to_numeric(df[val_col], errors='coerce')
    y_vals = sorted(df['Ypos'].dropna().unique(), reverse=True)
    x_vals = sorted(df['Xpos'].dropna().unique(), reverse=True)
    pivot = df.pivot_table(index="Ypos", columns="Xpos", values=val_col)
    pivot = pivot.reindex(index=y_vals, columns=x_vals)
    return pivot

def write_statistics(ws, row, col, cell_range):
    stats = {
        "count": f"=COUNT({cell_range})",
        "min": f"=MIN({cell_range})",
        "max": f"=MAX({cell_range})",
        "mean": f"=AVERAGE({cell_range})",
        "median": f"=MEDIAN({cell_range})",
        "stdev": f"=STDEV({cell_range})",
    }
    for idx, (key, formula) in enumerate(stats.items()):
        ws.cell(row=row + idx, column=col, value=key).alignment = center_align
        cell = ws.cell(row=row + idx, column=col + 1, value=formula)
        cell.alignment = center_align
        ws.cell(row=row + idx, column=col).border = ws.cell(row=row + idx, column=col + 1).border = thin_border

def format_numeric_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "NaN":
                cell.value = None
            elif isinstance(cell.value, float):
                if cell.value.is_integer():
                    cell.value = int(cell.value)
                else:
                    try:
                        decimal_str = f"{cell.value:.10f}".rstrip("0").split(".")[1]
                        if len(decimal_str) > 4:
                            cell.value = float(f"{cell.value:.2E}")
                            cell.number_format = "0.00E+00"
                    except (IndexError, ValueError):
                        continue

def analyze_excel_by_lot(excel_file_path, lot_column_name='Lot', start_column_index=6):
    """
    整片 wafer LOT 統計分析：
    - 自動清理 'Over'、'N/A'、'null' 等非數值資料
    - 對每個工作表依 LOT 計算 min, max, mean, median, stdev, count
    - 轉置後輸出 _統計.xlsx
    """
    xls = pd.ExcelFile(excel_file_path)
    all_sheets_statistics = {}

    for sheet_name in xls.sheet_names:
        df = xls.parse(sheet_name)
        if lot_column_name not in df.columns:
            continue

        df = df.replace(
            to_replace=[
                "Over", "OVER", "over",
                "N/A", "n/a", "NA", "na",
                "null", "NULL", "-", "—", " ", ""
            ],
            value=np.nan
        )

        data_columns = df.columns[start_column_index:].tolist()
        if not data_columns:
            continue

        for col in data_columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

        valid_cols = [col for col in data_columns if df[col].notna().any()]
        if not valid_cols:
            continue

        agg_funcs = {
            'min': 'min',
            'max': 'max',
            'mean': 'mean',
            'median': 'median',
            'stdev': 'std',
            'counts': 'count'
        }
        agg_dict = {col: list(agg_funcs.values()) for col in valid_cols}

        grouped_stats = df.groupby(lot_column_name).agg(agg_dict)

        transposed_stats = grouped_stats.T
        new_index_names = [f"{stat}_{item}" for item, stat in transposed_stats.index]
        transposed_stats.index = new_index_names
        transposed_stats = transposed_stats.sort_index()

        all_sheets_statistics[sheet_name] = transposed_stats

    stat_path = excel_file_path.replace('.xlsx', '_統計.xlsx')
    with pd.ExcelWriter(stat_path, engine='openpyxl') as writer:
        for sheet_name, stats_df in all_sheets_statistics.items():
            stats_df.to_excel(writer, sheet_name=f"{sheet_name}_統計")

    print(f"✅ 整片 wafer 統計完成（含 Over→空值 清理）：{stat_path}")


# =============================
# 主程式執行區段（整合）
# =============================
if __name__ == "__main__":
    root = Tk()
    root.withdraw()

    folder_in = filedialog.askdirectory(title="📂 選擇來源資料夾")
    if not folder_in:
        print("❌ 未選擇來源資料夾")
        exit()

    folder_out = filedialog.askdirectory(title="📁 選擇輸出資料夾")
    if not folder_out:
        print("❌ 未選擇輸出資料夾")
        exit()

    header_row = simpledialog.askinteger("欄位列選擇", "請輸入標題所在的列號（例如13）", initialvalue=13)
    if header_row is None:
        print("❌ 未輸入標題列號")
        exit()

    merged_path = merge_excel_with_custom_header(folder_in, folder_out, header_row_index=header_row - 1)
    if not merged_path:
        exit()

    time.sleep(2)  # 等待檔案穩定寫入
    df = pd.read_excel(merged_path)

    value_columns = [
        'IR_1V(uA)',
        'IR_10V(uA)',
        'IR_30V(uA)',
        'IR_50V(uA)',
        'IR_60V(uA)',
        'IR_68V(uA)',
        'IR_71V(uA)',
        'IR_74V(uA)',
        'VR_10uA(V)',
        'VR_100uA(V)',
        'VR_1mA(V)',
        'VR_10mA(V)',
        'VF_10mA(V)',
        'VF_260mA(V)',
        'VF_1A(V)',
        'VF_10A(V)'
    ]

    # =============================
    # MAP 畫圖 + 統計（整片 wafer）
    # =============================
    wb = Workbook()
    ws = wb.create_sheet(title="全片")

    lot_groups = {}
    for _, row in df.iterrows():
        lot = str(row["Lot"])
        if "_" in lot:
            main_lot, sub_lot = lot.rsplit("_", 1)
            lot_groups.setdefault(main_lot, {}).setdefault(sub_lot, []).append(row)
        else:
            lot_groups.setdefault(lot, {}).setdefault("", []).append(row)

    group_idx, row_offset, col_offset = 0, 0, 0
    max_groups_per_row = 10

    for main_lot, sublots in lot_groups.items():
        for sub_lot, rows in sublots.items():
            lot_df = pd.DataFrame(rows)
            lot_label = f"{main_lot}_{sub_lot}" if sub_lot else main_lot
            for val_col in [col for col in value_columns if col in lot_df.columns]:
                pivot = build_pivot(lot_df, val_col)
                nrows, ncols = pivot.shape
                base_row, base_col = row_offset + 1, col_offset + 1
                label = f"{lot_label}-{val_col}"
                ws.cell(row=base_row, column=base_col, value=label).alignment = left_align
                for i, row_vals in enumerate(pivot.values, start=1):
                    for j, val in enumerate(row_vals, start=1):
                        cell = ws.cell(row=base_row + i, column=base_col + j - 1,
                                       value="NaN" if pd.isna(val) else val)
                        cell.alignment = center_align
                        cell.border = thin_border
                if nrows > 0 and ncols > 0:
                    start_cell = ws.cell(row=base_row + 1, column=base_col).coordinate
                    end_cell = ws.cell(row=base_row + nrows, column=base_col + ncols - 1).coordinate
                    data_range = f"{start_cell}:{end_cell}"
                    rule = ColorScaleRule(start_type='min', start_color='FFFFFF',
                                         mid_type='percentile', mid_value=50,
                                         mid_color='FFFF00', end_type='max', end_color='FF0000')
                    ws.conditional_formatting.add(data_range, rule)
                    write_statistics(ws, base_row + nrows + 2, base_col, data_range)
                group_idx += 1
                if group_idx % max_groups_per_row == 0:
                    col_offset = 0
                    row_offset += (nrows + 10)
                else:
                    col_offset += (ncols + 2)

    # =============================
    # 常態分布圖（Excel 原生圖表）
    # 同一 wafer → 同一列；不同欄位 → 同一列往右排
    # =============================
    build_excel_dist_charts(wb, lot_groups, value_columns)

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    for ws in wb.worksheets:
        if ws.title != "分布圖":
            format_numeric_cells(ws)
    map_path = merged_path.replace('.xlsx', '_MAP.xlsx')
    wb.save(map_path)
    print(f"✅ 整片 wafer MAP 圖與統計已完成：{map_path}")

    # =============================
    # 整片 wafer 統計輸出
    # =============================
    analyze_excel_by_lot(merged_path, lot_column_name='Lot', start_column_index=6)
